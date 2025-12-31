import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from scipy import stats
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler

st.set_page_config(page_title="Doğalgaz Kaçak Tespit", layout="wide", page_icon="🔥")

# Başlık
st.title("🔥 Gelişmiş Doğalgaz Kaçak Kullanım Tespit Sistemi")
st.markdown("### 🤖 Makine Öğrenmesi ve İstatistiksel Analiz ile Anomali Tespiti")
st.markdown("---")

# Sidebar - Parametreler
with st.sidebar:
    st.header("⚙️ Analiz Parametreleri")
    
    st.subheader("📊 Temel Kriterler")
    dusus_esigi = st.slider("Ani Düşüş Eşiği (%)", 30, 95, 60, 5)
    sifir_ay = st.slider("Min. Sıfır Tüketim (Ay)", 1, 12, 3)
    bina_sapma_carpan = st.slider("Bina Z-Score Eşiği", 1.0, 4.0, 2.0, 0.5)
    min_bina_daire = st.number_input("Min. Daire Sayısı", 2, 20, 3)
    
    st.markdown("---")
    st.subheader("🧠 Makine Öğrenmesi")
    use_ml = st.checkbox("ML Anomali Tespiti Kullan", value=True)
    ml_contamination = st.slider("ML Anomali Oranı", 0.01, 0.20, 0.05, 0.01,
                                   help="Veri setindeki beklenen anomali oranı")
    
    st.markdown("---")
    st.subheader("📈 Trend Analizi")
    check_trend = st.checkbox("Trend Değişimi Analizi", value=True)
    trend_change_threshold = st.slider("Trend Değişim Eşiği (%)", 30, 90, 50, 10)
    
    st.markdown("---")
    st.subheader("🔍 Patern Analizi")
    check_seasonality = st.checkbox("Mevsimsellik Analizi", value=True)
    check_outliers = st.checkbox("İstatistiksel Aykırı Değerler", value=True)
    
    st.markdown("---")
    st.markdown("### 📋 Tespit Yöntemleri")
    st.markdown("""
    **1. Bina Karşılaştırma**
    - Z-score ile istatistiksel sapma
    - Binadaki diğer dairelerle karşılaştırma
    
    **2. Trend Analizi**
    - Tüketim trendinde ani değişim
    - Düşüş/artış paternleri
    
    **3. Makine Öğrenmesi**
    - Isolation Forest algoritması
    - Çok boyutlu anomali tespiti
    
    **4. İstatistiksel Testler**
    - Grubbs testi (aykırı değer)
    - Mevsimsellik kontrolü
    
    **5. Sıfır Tüketim**
    - Uzun süreli sıfır kayıtlar
    - Ardışık sıfır dönemler
    """)

# Dosya yükleme
uploaded_file = st.file_uploader("📁 Excel Dosyası Yükleyin", type=['xlsx', 'xls'])

def calculate_trend(values):
    """Lineer trend hesapla"""
    x = np.arange(len(values))
    valid_idx = ~np.isnan(values)
    if np.sum(valid_idx) < 2:
        return 0, 0
    slope, intercept = np.polyfit(x[valid_idx], values[valid_idx], 1)
    return slope, intercept

def detect_trend_change(values, window=6):
    """Trend değişimi tespit et"""
    if len(values) < window * 2:
        return []
    
    changes = []
    for i in range(window, len(values) - window):
        before = values[i-window:i]
        after = values[i:i+window]
        
        if len(before[before > 0]) < 3 or len(after[after > 0]) < 3:
            continue
        
        slope_before, _ = calculate_trend(before)
        slope_after, _ = calculate_trend(after)
        
        if slope_before != 0:
            change_pct = abs((slope_after - slope_before) / slope_before * 100)
            if change_pct > trend_change_threshold:
                changes.append({
                    'index': i,
                    'slope_before': slope_before,
                    'slope_after': slope_after,
                    'change_pct': change_pct
                })
    
    return changes

def grubbs_test(data, alpha=0.05):
    """Grubbs testi ile aykırı değer tespiti"""
    data = data[data > 0]
    if len(data) < 3:
        return []
    
    outliers = []
    while True:
        mean = np.mean(data)
        std = np.std(data)
        if std == 0:
            break
        
        abs_val = np.abs(data - mean)
        max_idx = np.argmax(abs_val)
        max_val = data[max_idx]
        G = abs_val[max_idx] / std
        
        n = len(data)
        t_dist = stats.t.ppf(1 - alpha / (2 * n), n - 2)
        threshold = ((n - 1) * np.sqrt(np.square(t_dist))) / (np.sqrt(n) * np.sqrt(n - 2 + np.square(t_dist)))
        
        if G > threshold:
            outliers.append(max_val)
            data = np.delete(data, max_idx)
        else:
            break
        
        if len(data) < 3:
            break
    
    return outliers

def check_seasonality(values, period=12):
    """Mevsimsellik kontrolü - basit yöntem"""
    if len(values) < period * 2:
        return False, 0
    
    values = values[values > 0]
    if len(values) < period:
        return False, 0
    
    # Otokorelasyon hesapla
    mean = np.mean(values)
    var = np.var(values)
    if var == 0:
        return False, 0
    
    autocorr = np.correlate(values - mean, values - mean, mode='full')
    autocorr = autocorr[len(autocorr)//2:]
    autocorr = autocorr / (var * len(values))
    
    if len(autocorr) > period:
        seasonal_corr = autocorr[period]
        return seasonal_corr > 0.3, seasonal_corr
    
    return False, 0

if uploaded_file is not None:
    try:
        # Excel dosyasını oku
        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip()
        
        st.success(f"✅ {len(df)} satır veri yüklendi")
        
        # Ay sütunlarını bul
        ay_sutunlari = [col for col in df.columns if '/' in str(col) or (col not in ['tn', 'bn'] and col.replace('.','').isdigit())]
        
        # Veriyi numerik yap
        for col in ay_sutunlari:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        st.markdown("---")
        
        # Genel istatistikler
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Toplam Tesisat", len(df))
        with col2:
            st.metric("Toplam Bina", df['bn'].nunique())
        with col3:
            st.metric("Analiz Ayı", len(ay_sutunlari))
        with col4:
            toplam_tuketim = df[ay_sutunlari].sum().sum()
            st.metric("Toplam Tüketim", f"{toplam_tuketim:,.0f}")
        
        st.markdown("---")
        
        # Analiz başlat
        with st.spinner("🔍 Gelişmiş anomali tespiti yapılıyor..."):
            
            all_anomalies = {}  # Tesisat bazında tüm anomaliler
            
            # 1. Bina bazlı analiz
            st.info("📊 1/6 - Bina bazlı istatistiksel analiz...")
            bina_anomaliler = []
            
            for bina in df['bn'].unique():
                bina_df = df[df['bn'] == bina].copy()
                
                if len(bina_df) < min_bina_daire:
                    continue
                
                for ay in ay_sutunlari:
                    bina_values = bina_df[ay].values
                    bina_ort = np.mean(bina_values)
                    bina_std = np.std(bina_values)
                    
                    if bina_std == 0 or pd.isna(bina_std) or bina_ort < 5:
                        continue
                    
                    for idx, row in bina_df.iterrows():
                        deger = row[ay]
                        z_score = (deger - bina_ort) / bina_std
                        
                        if abs(z_score) > bina_sapma_carpan:
                            tn = row['tn']
                            if tn not in all_anomalies:
                                all_anomalies[tn] = {'tn': tn, 'bn': row['bn'], 'anomalies': []}
                            
                            all_anomalies[tn]['anomalies'].append({
                                'type': 'Bina Anomalisi',
                                'ay': ay,
                                'deger': deger,
                                'bina_ort': bina_ort,
                                'z_score': z_score,
                                'severity': 'high' if abs(z_score) > 3 else 'medium'
                            })
            
            # 2. Ani düşüş analizi
            st.info("📉 2/6 - Ani düşüş ve değişim analizi...")
            for idx, row in df.iterrows():
                tn = row['tn']
                bn = row['bn']
                
                for i in range(1, len(ay_sutunlari)):
                    onceki = row[ay_sutunlari[i-1]]
                    simdiki = row[ay_sutunlari[i]]
                    
                    if onceki > 10 and simdiki >= 0:
                        dusus_orani = ((onceki - simdiki) / onceki) * 100
                        
                        if dusus_orani >= dusus_esigi:
                            if tn not in all_anomalies:
                                all_anomalies[tn] = {'tn': tn, 'bn': bn, 'anomalies': []}
                            
                            all_anomalies[tn]['anomalies'].append({
                                'type': 'Ani Düşüş',
                                'ay': ay_sutunlari[i],
                                'onceki_ay': ay_sutunlari[i-1],
                                'onceki_deger': onceki,
                                'deger': simdiki,
                                'dusus_orani': dusus_orani,
                                'severity': 'high' if dusus_orani > 80 else 'medium'
                            })
            
            # 3. Sıfır tüketim analizi
            st.info("⭕ 3/6 - Sıfır tüketim dönemleri analizi...")
            for idx, row in df.iterrows():
                tn = row['tn']
                bn = row['bn']
                sifir_sayaci = 0
                baslangic = None
                
                for ay in ay_sutunlari:
                    if row[ay] == 0:
                        if sifir_sayaci == 0:
                            baslangic = ay
                        sifir_sayaci += 1
                    else:
                        if sifir_sayaci >= sifir_ay:
                            if tn not in all_anomalies:
                                all_anomalies[tn] = {'tn': tn, 'bn': bn, 'anomalies': []}
                            
                            all_anomalies[tn]['anomalies'].append({
                                'type': 'Sıfır Tüketim',
                                'baslangic': baslangic,
                                'bitis': ay_sutunlari[ay_sutunlari.index(ay) - 1],
                                'sure_ay': sifir_sayaci,
                                'severity': 'high' if sifir_sayaci >= 6 else 'medium'
                            })
                        sifir_sayaci = 0
                        baslangic = None
                
                if sifir_sayaci >= sifir_ay:
                    if tn not in all_anomalies:
                        all_anomalies[tn] = {'tn': tn, 'bn': bn, 'anomalies': []}
                    
                    all_anomalies[tn]['anomalies'].append({
                        'type': 'Sıfır Tüketim',
                        'baslangic': baslangic,
                        'bitis': ay_sutunlari[-1],
                        'sure_ay': sifir_sayaci,
                        'severity': 'high' if sifir_sayaci >= 6 else 'medium'
                    })
            
            # 4. Trend değişimi analizi
            if check_trend:
                st.info("📈 4/6 - Trend değişimi analizi...")
                for idx, row in df.iterrows():
                    tn = row['tn']
                    bn = row['bn']
                    values = row[ay_sutunlari].values
                    
                    trend_changes = detect_trend_change(values)
                    
                    if trend_changes:
                        if tn not in all_anomalies:
                            all_anomalies[tn] = {'tn': tn, 'bn': bn, 'anomalies': []}
                        
                        for tc in trend_changes:
                            all_anomalies[tn]['anomalies'].append({
                                'type': 'Trend Değişimi',
                                'ay': ay_sutunlari[tc['index']],
                                'degisim_orani': tc['change_pct'],
                                'onceki_trend': tc['slope_before'],
                                'sonraki_trend': tc['slope_after'],
                                'severity': 'high' if tc['change_pct'] > 80 else 'medium'
                            })
            
            # 5. İstatistiksel aykırı değer (Grubbs test)
            if check_outliers:
                st.info("🔬 5/6 - İstatistiksel aykırı değer analizi...")
                for idx, row in df.iterrows():
                    tn = row['tn']
                    bn = row['bn']
                    values = row[ay_sutunlari].values
                    
                    outliers = grubbs_test(values.copy())
                    
                    if outliers:
                        if tn not in all_anomalies:
                            all_anomalies[tn] = {'tn': tn, 'bn': bn, 'anomalies': []}
                        
                        all_anomalies[tn]['anomalies'].append({
                            'type': 'İstatistiksel Aykırı Değer',
                            'outlier_count': len(outliers),
                            'outlier_values': outliers,
                            'severity': 'medium'
                        })
            
            # 6. Machine Learning - Isolation Forest
            if use_ml:
                st.info("🤖 6/6 - Makine öğrenmesi anomali tespiti...")
                
                # Özellik mühendisliği
                features_list = []
                tn_list = []
                
                for idx, row in df.iterrows():
                    values = row[ay_sutunlari].values
                    non_zero = values[values > 0]
                    
                    if len(non_zero) < 3:
                        continue
                    
                    features = {
                        'mean': np.mean(non_zero),
                        'std': np.std(non_zero),
                        'cv': (np.std(non_zero) / np.mean(non_zero)) if np.mean(non_zero) > 0 else 0,
                        'max': np.max(values),
                        'min': np.min(non_zero),
                        'range': np.max(values) - np.min(non_zero),
                        'zero_count': np.sum(values == 0),
                        'trend': calculate_trend(values)[0],
                        'q1': np.percentile(non_zero, 25),
                        'q3': np.percentile(non_zero, 75),
                    }
                    
                    features_list.append(list(features.values()))
                    tn_list.append(row['tn'])
                
                if len(features_list) > 10:
                    X = np.array(features_list)
                    
                    # Normalizasyon
                    scaler = StandardScaler()
                    X_scaled = scaler.fit_transform(X)
                    
                    # Isolation Forest
                    iso_forest = IsolationForest(contamination=ml_contamination, random_state=42)
                    predictions = iso_forest.fit_predict(X_scaled)
                    scores = iso_forest.score_samples(X_scaled)
                    
                    # Anomali olanları işaretle
                    for i, (pred, score) in enumerate(zip(predictions, scores)):
                        if pred == -1:  # Anomali
                            tn = tn_list[i]
                            bn = df[df['tn'] == tn]['bn'].values[0]
                            
                            if tn not in all_anomalies:
                                all_anomalies[tn] = {'tn': tn, 'bn': bn, 'anomalies': []}
                            
                            all_anomalies[tn]['anomalies'].append({
                                'type': 'ML Anomali',
                                'anomaly_score': abs(score),
                                'severity': 'high' if abs(score) > 0.5 else 'medium'
                            })
        
        # Risk skoru hesapla
        def calculate_comprehensive_risk(anomalies):
            score = 0
            weights = {
                'Bina Anomalisi': 35,
                'Ani Düşüş': 30,
                'Sıfır Tüketim': 25,
                'Trend Değişimi': 20,
                'ML Anomali': 40,
                'İstatistiksel Aykırı Değer': 15
            }
            
            severity_multiplier = {'high': 1.5, 'medium': 1.0, 'low': 0.5}
            
            for anom in anomalies:
                base_score = weights.get(anom['type'], 10)
                mult = severity_multiplier.get(anom.get('severity', 'medium'), 1.0)
                score += base_score * mult
            
            return score
        
        # Sonuçları hazırla
        results = []
        for tn, data in all_anomalies.items():
            risk_score = calculate_comprehensive_risk(data['anomalies'])
            
            # Anomali sayıları
            anom_counts = {}
            for anom in data['anomalies']:
                anom_type = anom['type']
                anom_counts[anom_type] = anom_counts.get(anom_type, 0) + 1
            
            results.append({
                'tn': tn,
                'bn': data['bn'],
                'risk_score': risk_score,
                'anomaly_count': len(data['anomalies']),
                'anomaly_types': anom_counts,
                'anomalies': data['anomalies']
            })
        
        # Risk skoruna göre sırala
        results.sort(key=lambda x: x['risk_score'], reverse=True)
        
        # Sonuçlar
        st.success("✅ Analiz tamamlandı!")
        st.markdown("---")
        st.header("📊 Analiz Sonuçları")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("🚨 Toplam Şüpheli", len(results))
        with col2:
            high_risk = sum(1 for r in results if r['risk_score'] >= 150)
            st.metric("🔴 Yüksek Risk", high_risk)
        with col3:
            medium_risk = sum(1 for r in results if 80 <= r['risk_score'] < 150)
            st.metric("🟡 Orta Risk", medium_risk)
        with col4:
            low_risk = sum(1 for r in results if r['risk_score'] < 80)
            st.metric("🟢 Düşük Risk", low_risk)
        
        if results:
            st.markdown("---")
            st.subheader(f"🔍 En Yüksek Riskli {min(20, len(results))} Tesisat")
            
            # Detaylı sonuçlar
            for result in results[:20]:
                tn = result['tn']
                bn = result['bn']
                score = result['risk_score']
                
                # Risk seviyesi
                if score >= 150:
                    risk_color = "🔴"
                    risk_label = "KRİTİK RİSK"
                    border_color = "red"
                elif score >= 80:
                    risk_color = "🟡"
                    risk_label = "ORTA RİSK"
                    border_color = "orange"
                else:
                    risk_color = "🟢"
                    risk_label = "DÜŞÜK RİSK"
                    border_color = "green"
                
                with st.expander(f"{risk_color} **Tesisat: {tn}** | Bina: {bn} | Risk Skoru: {score:.1f} - {risk_label}"):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        # Tüketim grafiği + bina ortalaması
                        tesisat_data = df[df['tn'] == tn][ay_sutunlari].values[0]
                        bina_data = df[df['bn'] == bn][ay_sutunlari].mean().values
                        
                        fig = go.Figure()
                        fig.add_trace(go.Scatter(
                            x=ay_sutunlari, y=tesisat_data,
                            name='Tesisat', mode='lines+markers',
                            line=dict(color='red', width=3),
                            marker=dict(size=8)
                        ))
                        fig.add_trace(go.Scatter(
                            x=ay_sutunlari, y=bina_data,
                            name='Bina Ortalaması', mode='lines',
                            line=dict(color='blue', width=2, dash='dash')
                        ))
                        
                        # Anomali noktalarını işaretle
                        anomaly_months = []
                        anomaly_values = []
                        for anom in result['anomalies']:
                            if 'ay' in anom and anom['ay'] in ay_sutunlari:
                                idx = ay_sutunlari.index(anom['ay'])
                                anomaly_months.append(anom['ay'])
                                anomaly_values.append(tesisat_data[idx])
                        
                        if anomaly_months:
                            fig.add_trace(go.Scatter(
                                x=anomaly_months, y=anomaly_values,
                                name='Anomali', mode='markers',
                                marker=dict(color='orange', size=15, symbol='x', line=dict(width=2))
                            ))
                        
                        fig.update_layout(
                            title=f'Tesisat {tn} - Bina {bn} Tüketim Analizi',
                            xaxis_title='Ay',
                            yaxis_title='Tüketim',
                            height=350,
                            hovermode='x unified',
                            showlegend=True
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    
                    with col2:
                        st.markdown("### 📋 Anomali Detayları")
                        st.markdown(f"**Toplam Anomali:** {result['anomaly_count']}")
                        
                        # Anomali tiplerine göre grupla
                        for anom_type, count in result['anomaly_types'].items():
                            st.markdown(f"**{anom_type}:** {count}")
                        
                        st.markdown("---")
                        st.markdown("### 🔎 Detaylar")
                        
                        for anom in result['anomalies'][:5]:
                            anom_type = anom['type']
                            severity_emoji = "🔴" if anom.get('severity') == 'high' else "🟡"
                            
                            st.markdown(f"{severity_emoji} **{anom_type}**")
                            
                            if anom_type == 'Bina Anomalisi':
                                st.markdown(f"- {anom['ay']}: {anom['deger']:.0f} (Z-score: {anom['z_score']:.2f})")
                                st.markdown(f"  Bina ort: {anom['bina_ort']:.0f}")
                            
                            elif anom_type == 'Ani Düşüş':
                                st.markdown(f"- {anom['ay']}: %{anom['dusus_orani']:.1f} düşüş")
                                st.markdown(f"  {anom['onceki_deger']:.0f} → {anom['deger']:.0f}")
                            
                            elif anom_type == 'Sıfır Tüketim':
                                st.markdown(f"- {anom['baslangic']} - {anom['bitis']}")
                                st.markdown(f"  {anom['sure_ay']} ay sıfır")
                            
                            elif anom_type == 'Trend Değişimi':
                                st.markdown(f"- {anom['ay']}: %{anom['degisim_orani']:.1f} değişim")
                            
                            elif anom_type == 'ML Anomali':
                                st.markdown(f"- Anomali skoru: {anom['anomaly_score']:.3f}")
            
            # Excel raporu oluştur
            st.markdown("---")
            st.subheader("📥 Detaylı Rapor İndir")
            
            def create_detailed_excel(results, df, ay_sutunlari):
                output = BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Sheet 1: Özet
                    summary_data = []
                    for r in results:
                        row = {
                            'Tesisat No': r['tn'],
                            'Bina No': r['bn'],
                            'Risk Skoru': round(r['risk_score'], 2),
                            'Toplam Anomali': r['anomaly_count']
                        }
                        
                        # Risk seviyesi
                        if r['risk_score'] >= 150:
                            row['Risk Seviyesi'] = 'KRİTİK'
                        elif r['risk_score'] >= 80:
                            row['Risk Seviyesi'] = 'ORTA'
                        else:
                            row['Risk Seviyesi'] = 'DÜŞÜK'
                        
                        # Anomali tipleri
                        for anom_type, count in r['anomaly_types'].items():
                            row[anom_type] = count
                        
                        # Aylık veriler
                        tesisat_row = df[df['tn'] == r['tn']].iloc[0]
                        for ay in ay_sutunlari:
                            row[ay] = tesisat_row[ay]
                        
                        summary_data.append(row)
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Özet Rapor', index=False)
                    
                    # Sheet 2: Detaylı anomaliler
                    detailed_data = []
                    for r in results:
                        for anom in r['anomalies']:
                            detailed_row = {
                                'Tesisat No': r['tn'],
                                'Bina No': r['bn'],
                                'Anomali Tipi': anom['type'],
                                'Önem': anom.get('severity', 'medium').upper()
                            }
                            
                            # Tip-specific detaylar
                            if 'ay' in anom:
                                detailed_row['Ay'] = anom['ay']
                            if 'deger' in anom:
                                detailed_row['Değer'] = anom['deger']
                            if 'dusus_orani' in anom:
                                detailed_row['Düşüş Oranı (%)'] = round(anom['dusus_orani'], 2)
                            if 'z_score' in anom:
                                detailed_row['Z-Score'] = round(anom['z_score'], 2)
                            if 'sure_ay' in anom:
                                detailed_row['Süre (Ay)'] = anom['sure_ay']
                            
                            detailed_data.append(detailed_row)
                    
                    if detailed_data:
                        detailed_df = pd.DataFrame(detailed_data)
                        detailed_df.to_excel(writer, sheet_name='Detaylı Anomaliler', index=False)
                
                output.seek(0)
                
                # Stil ekle
                wb = openpyxl.load_workbook(output)
                
                # Özet rapor stil
                ws = wb['Özet Rapor']
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Risk seviyesine göre renklendirme
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                
                risk_col = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Risk Seviyesi':
                        risk_col = idx
                        break
                
                if risk_col:
                    for row in range(2, ws.max_row + 1):
                        risk_val = ws.cell(row=row, column=risk_col).value
                        if risk_val == 'KRİTİK':
                            fill = red_fill
                        elif risk_val == 'ORTA':
                            fill = yellow_fill
                        else:
                            fill = green_fill
                        
                        for col in range(1, 8):  # İlk 7 sütunu renklendir
                            ws.cell(row=row, column=col).fill = fill
                
                # Sütun genişlikleri
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
                
                # Detaylı anomaliler sayfası stil
                if 'Detaylı Anomaliler' in wb.sheetnames:
                    ws2 = wb['Detaylı Anomaliler']
                    for cell in ws2[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    for column in ws2.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        ws2.column_dimensions[column_letter].width = min(max_length + 2, 25)
                
                output2 = BytesIO()
                wb.save(output2)
                output2.seek(0)
                
                return output2.getvalue()
            
            excel_data = create_detailed_excel(results, df, ay_sutunlari)
            
            st.download_button(
                label="📊 Detaylı Excel Raporu İndir (Tüm Anomaliler)",
                data=excel_data,
                file_name=f"dogalgaz_anomali_raporu_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        else:
            st.success("✅ Belirlenen kriterlere göre şüpheli tesisat bulunamadı!")
            st.info("Parametreleri gevşeterek daha fazla anomali tespit edebilirsiniz.")
    
    except Exception as e:
        st.error(f"❌ Hata: {str(e)}")
        import traceback
        st.code(traceback.format_exc())

else:
    st.info("👆 Lütfen yukarıdan Excel dosyanızı yükleyin.")
    
    st.markdown("---")
    
    # Örnek veri göster
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📝 Beklenen Dosya Formatı")
        st.markdown("""
        | tn | bn | 2023/07 | 2023/08 | 2023/09 | ... |
        |----|----|---------|---------|---------| --- |
        | 10009832 | 100003724 | 18.49 | 18.43 | 8.18 | ... |
        | 10009992 | 100003724 | 25.51 | 26.40 | 13.78 | ... |
        """)
    
    with col2:
        st.markdown("### 🎯 Tespit Edilen Anomali Tipleri")
        st.markdown("""
        1. **Bina Anomalisi**: Binadaki diğer dairelerden istatistiksel sapma
        2. **Ani Düşüş**: Keskin tüketim düşüşleri
        3. **Sıfır Tüketim**: Uzun süre sıfır kayıt
        4. **Trend Değişimi**: Tüketim trendinde ani değişim
        5. **ML Anomali**: Makine öğrenmesi tespiti
        6. **İstatistiksel Aykırı**: Grubbs test ile tespit
        """)
